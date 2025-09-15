import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from utils.path_helper import resource_path

class FormatadorExcel:
    """
    Uma classe dedicada a centralizar toda a lógica de formatação
    de arquivos Excel para o projeto SAP-UFAC.
    """
    
    @staticmethod
    def formatar_planilha_desfazimento(workbook, sheet, nome_planilha, dados):
        """
        Aplica uma formatação padrão a uma planilha de desfazimento.
        """
        # 1. Limpa a aba antes de popular
        sheet.delete_rows(1, sheet.max_row)

        # 2. Adiciona e formata o título principal
        sheet.merge_cells('A1:H1')
        titulo_cell = sheet['A1']
        titulo_cell.value = nome_planilha
        titulo_cell.font = Font(bold=True, size=14, name='Times New Roman')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 3. Adiciona o cabeçalho
        cabecalho = ['Nº DE ORDEM', 'TOMBO', 'DESCRIÇÃO DO BEM', 'DATA DA AQUISIÇÃO', 'DOCUMENTO FISCAL', 'UNIDADE RESPONSÁVEL', 'CLASSIFICAÇÃO', 'DESTINAÇÃO']
        sheet.append(cabecalho)

        # 4. Adiciona os dados
        for linha in dados:
            sheet.append(linha)

        # 5. Define os estilos
        alinhamento_central_com_quebra = Alignment(horizontal='center', vertical='center', wrap_text=True)
        borda_fina = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        fonte_cabecalho = Font(bold=True, size=12, name='Times New Roman')
        fonte_dados = Font(size=12, name='Times New Roman')
        
        # 6. Define larguras
        larguras = {'A': 15, 'B': 15, 'C': 40, 'D': 20, 'E': 25, 'F': 40, 'G': 20, 'H': 20}
        for letra_coluna, largura in larguras.items():
            sheet.column_dimensions[letra_coluna].width = largura

        # 7. Formata o cabeçalho
        sheet.row_dimensions[2].height = 40
        for cell in sheet[2]:
            cell.alignment = alinhamento_central_com_quebra
            cell.border = borda_fina
            cell.font = fonte_cabecalho

        # 8. Formata os dados
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            sheet.row_dimensions[row[0].row].height = 30
            for cell in row:
                cell.alignment = alinhamento_central_com_quebra
                cell.border = borda_fina
                cell.font = fonte_dados